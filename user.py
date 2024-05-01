from re import T
import openpyxl 
import ttkbootstrap as tb


class user():
    def __init__(self, username):
        path = "database_offline.xlsx"
        self.wb_obj = openpyxl.load_workbook(path)
        self.wks_account = self.wb_obj["accountInfo"]
        self.wks_store = self.wb_obj["store"]

        self.username = username
 
    def find_row(self):
        """
            returns:
            the row the username belongs to (int)
        """
        name = ""
        empty_row = None
        row = 0
        row_count = self.wks_account.max_row
        for i in range(1, row_count + 1):
            name = self.wks_account.cell(i, 1).value
            if name == self.username:
                return i
            if name is None and empty_row is None:
                empty_row = i

        if empty_row is None:
            empty_row = row_count + 1

        self.wks_account.cell(empty_row, 1).value = self.username
        self.wb_obj.save(self.path)

        return empty_row

    def get_username(self):
        """
            returns:
            username (string)
        """
        
        return self.username

    def get_password(self):
        """
            returns:
            user's password (string)
        """
        password = self.wks_account.cell(self.find_row(), 2).value
        return password

    def get_email(self):
        """
            returns:
            user's email (string)
        """
        email = self.wks_account.cell(self.find_row(), 3).value
        return self.get_email

    def get_parsed_library(self):
        # Library section of database format ex: The Witcher 3: Wild Hunt/Civilization VI/Half-Life: Alyx/Command & Conquer/Doom
        # Gets library only as titles
        """
            returns:
            titles of games belonging to the user
        """
        library = self.wks_account.cell(self.find_row(),4).value
        if(library != None):
            parsed_library = library.split("/")
            return parsed_library
        return []

    def update_library(self, new):
        """
            args:
            title of a game (string)
            updates the users library by appending the string in the cell with the title a purchase
        
        """
        library = self.wks_account.cell(self.find_row(),4).value
        if(library != None):
            self.wks_account.cell(self.find_row(),4).value +=  "/" + new 
        else:
            self.wks_account.cell(self.find_row(),4).value = new
        self.wb_obj.save("database_offline.xlsx")
    
    def check_inlibrary(self, game):
        """
        Using game title checks if title is in library
        args:
            game (string)
        returns:
            status (string)

        """
        status = False
        titles = self.get_parsed_library()
        for title in titles:
            if(title == game):
                status = True
                break
        return status
 
    
    def get_game_details(self):
        #Get games in library in dictionary format
        """
            returns:
            user's library in a list with each game as a dictionary
        """

        titles = self.get_parsed_library()
        library = []
        count = 0 
        row_count = 31
        for title in titles:
            for i in range(1, row_count-1):
                cell_title = self.wks_store.cell(i+1, 2).value 
                if (title == cell_title):
                    library.append({})
                    library[count]["Title"] =  self.wks_store.cell(i+1, 2).value
                    library[count]["Developer"] =  self.wks_store.cell(i+1, 3).value
                    library[count]["Price"] =  self.wks_store.cell(i+1, 4).value
                    tags = self.wks_store.cell(i+1, 5).value
                    parsed_tags = tags.split(",")
                    library[count]["Tags"] =  parsed_tags
                    library[count]["Release_Date"] = self.wks_store.cell(i+1, 6).value
                    count += 1
        return library
    
    
    def getParsedFriendsList(self):
        # Get friends list associated with user account
        friendsList = self.wks_account.cell(self.find_row(), 5).value
        if (friendsList != None):
            parsedFriendsList = friendsList.split(">>")
            return parsedFriendsList
        return []
    
    def updateFriendsList(self, newFriend):
        # Add new friend to account data
        friendsList = self.wks_account.cell(self.find_row(), 5).value
        if (friendsList != None):
            self.wks_account.cell(self.find_row(), 5).value += '>>' + newFriend
        else:
            self.wks_account.cell(self.find_row(), 5).value = newFriend
        
        self.wb_obj.save("database_offline.xlsx")
    
   
    def sendFriendRequest(self, friendUsername):
        # sends friend request and adds it into the other user's database
        friendVar = tb.StringVar(value=friendUsername)
        friend = user(friendVar.get())
        friendIdLoc = self.wks_account.cell(friend.find_row(), 1).row
        message = "{profile} has sent you a friend request.".format(profile = self.get_username())
        friendRequests = self.wks_account.cell(friendIdLoc, 6).value 
        if (friendRequests != None):
            self.wks_account.cell(friendIdLoc, 6).value += ">>" + message
        else:
            self.wks_account.cell(friendIdLoc, 6).value = message
        
        self.wb_obj.save("database_offline.xlsx")


    def listFriendRequests(self):
        # display friend requests awaiting response
        friendRequests = self.wks_account.cell(self.find_row(), 6).value
        if (friendRequests != None):
            parseFriendRequests = friendRequests.split(">>")
            return parseFriendRequests
        
        return []
                 

    