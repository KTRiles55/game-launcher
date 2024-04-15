import openpyxl 

class store_off():
    def __init__(self):
        path = "database_offline.xlsx"
        wb_obj = openpyxl.load_workbook(path)
        self.wks = wb_obj["store"]
        
    def get_all_games(self):
        #return list of games with the same category 
        games = []
        count = 0 
        col_count = 31
        for i in range(1, col_count-1):
            games.append({})
            games[i-1]["Title"] =   self.wks.cell(i+1, 2).value
            games[i-1]["Developer"] =  self.wks.cell(i+1, 3).value
            games[i-1]["Price"] =  self.wks.cell(i+1, 4).value
            tag = self.wks.cell(i+1, 5).value
            parsed_tags = tag.split(",")
            games[i-1]["Tags"] =  parsed_tags

        return games

    def get_games_sharing_tag(self, selected):
        #return list of games with the same category 
        games = []
        count = 0 
        col_count = 31
        for i in range(1, col_count-1):
            tags = self.wks.cell(i+1, 5).value
            if(selected in tags):
                games.append({})
                games[count]["Title"] =  self.wks.cell(i+1, 2).value
                games[count]["Developer"] =  self.wks.cell(i+1, 3).value
                games[count]["Price"] =  self.wks.cell(i+1, 4).value
                parsed_tags = tags.split(",")
                games[count]["Tags"] =  parsed_tags
                count += 1

        return games
    
    def get_game(self,input):
        #gets specific game based on searchbar
        pass

    def get_all_titles(self):
        titles = []
        col_count = 31
        for i in range(1, col_count-1):
            titles.append(self.wks.cell(i+1, 2).value)

        return titles

    def get_related_search(self, entry):
        games = []
        count = 0
        col_count = 31
        for i in range(1, col_count-1):
            title = self.wks.cell(i+1, 2).value
            if entry.lower() in title.lower():
                tags = self.wks.cell(i+1, 5).value
                games.append({})
                games[count]["Title"] =  self.wks.cell(i+1, 2).value
                games[count]["Developer"] =  self.wks.cell(i+1, 3).value
                games[count]["Price"] =  self.wks.cell(i+1, 4).value
                parsed_tags = tags.split(",")
                games[count]["Tags"] =  parsed_tags
                count += 1 
        return games



