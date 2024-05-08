import openpyxl 
import string
import random

class store():
    def __init__(self):
        self.path = "database_offline.xlsx"
        self.wb_obj = openpyxl.load_workbook(self.path)
        self.wks = self.wb_obj["store"]
        

    def get_all_games(self):
        """
            returns:
            all games in as a dictionary in a list
        """
        games = []
        count = 0 
        col_count = 31
        for i in range(1, col_count-1):
            games.append({})
            games[i-1]["ID"] = self.wks.cell(i+1, 1).value
            games[i-1]["Title"] =   self.wks.cell(i+1, 2).value
            games[i-1]["Developer"] =  self.wks.cell(i+1, 3).value
            games[i-1]["Price"] =  self.wks.cell(i+1, 4).value
            tag = self.wks.cell(i+1, 5).value
            parsed_tags = tag.split(",")
            games[i-1]["Tags"] =  parsed_tags
            games[i-1]["Release_Date"] =  self.wks.cell(i+1, 6).value

        return games

    def get_games_sharing_tag(self, selected):
        #return list of games with the same category 
        """
            args:
            a category for a game (string)
            returns:
            all games that have the category in a list
        """
        games = []
        count = 0 
        col_count = 31
        if(selected == "All"):
            return self.get_all_games()
    
        for i in range(1, col_count-1):
            tags = self.wks.cell(i+1, 5).value
            if(selected in tags):
                games.append({})
                games[count]["Title"] =  self.wks.cell(i+1, 2).value
                games[count]["Developer"] =  self.wks.cell(i+1, 3).value
                games[count]["Price"] =  self.wks.cell(i+1, 4).value
                parsed_tags = tags.split(",")
                games[count]["Tags"] =  parsed_tags
                games[count]["Release_Date"] = self.wks.cell(i+1, 6).value
                count += 1

        return games


    def get_all_titles(self):
        """
            returns:
            a list of all game titles
        """
        titles = []
        col_count = 31
        for i in range(1, col_count-1):
            titles.append(self.wks.cell(i+1, 2).value)

        return titles

    def get_related_search(self, entry):
        """
            args:
            value from search bar (string)
            returns:
            a list of all games with similar title as a dictionary 
        """
        games = []
        count = 0
        col_count = 31
        for i in range(1, col_count-1):
            title = self.wks.cell(i+1, 2).value
            if entry.lower() in title.lower():
                tags = self.wks.cell(i+1, 5).value
                games.append({})
                games[count]["Title"] =  self.wks.cell(i+1, 2).value
                games[count]["Developer"] =  self.wks.cell(i+1, 3).value
                games[count]["Price"] =  self.wks.cell(i+1, 4).value
                parsed_tags = tags.split(",")
                games[count]["Tags"] =  parsed_tags
                games[count]["Release_Date"] = self.wks.cell(i+1, 6).value
                count += 1 
        return games

    def get_game_inrow(self, row):
        #Skips the title of a column
        """
            args:
            row number for a game (int)
            returns:
            a dictionary of game and related details
        """
        game = {}
        game["Title"] =  self.wks.cell(row, 2).value
        game["Developer"] =  self.wks.cell(row, 3).value
        game["Price"] =  self.wks.cell(row, 4).value
        tags = self.wks.cell(row, 5).value
        parsed_tags = tags.split(",")
        game["Tags"] =  parsed_tags
        game["Release_Date"] = self.wks.cell(row, 6).value
        return game

    def append_codes(self, title, code):
        #Code format 0D5YE91/1245NMT
        """
            args:
            title and gift code are both strings

            updates the database with valid gift codes
            
        """
        #Refresh spreadsheet
        self.wb_obj = openpyxl.load_workbook(self.path, data_only=True)
        self.wks_store = self.wb_obj["store"]

        count = 0
        col_count = 31
        for i in range(1, col_count-1):
            #Title cell location
            title_cell = self.wks_store.cell(i+1, 2).value
            if(title == title_cell):
                code_cell = self.wks_store.cell(i+1, 7).value
                if(code_cell != None):
                    self.wks_store.cell(i+1, 7).value  += "/" + code 
                else:
                    self.wks_store.cell(i+1, 7).value =  code
                break
        self.wb_obj.save("database_offline.xlsx")
        

    def validate_gift(self,input):
        """
            args:
            gift code as string
            returns:
            a game's title/id
        """
        #Refresh spreadsheet
        self.wb_obj = openpyxl.load_workbook(self.path, data_only=True)
        self.wks_store = self.wb_obj["store"]

        count = 0
        row_count = 31
        row_found = 0
        valid = False
        #Traverse games to find valid code
        for i in range(1, row_count-1):
            codes = self.wks_store.cell(i+1, 7).value
            if(codes != None):
                parsed_codes = codes.split("/")
                #print("lst=" + str(parsed_codes))
                for j in range(len(parsed_codes)):
                    if(parsed_codes[j] == input):
                        gift = self.get_game_inrow(i+1)
                        parsed_codes.pop(j)
                        row_found = i
                        valid = True
                        break
            if(valid == True):
                break
        #removes code from being valid
        if(valid == True):
            self.wks_store.cell(row_found, 7).value = None
            new_val = ""
            for i in range(len(parsed_codes)):
                new_val += "/" + parsed_codes[i]
            self.wks_store.cell(row_found+1, 7).value = new_val
            self.wb_obj.save("database_offline.xlsx")
            return gift
        return None

